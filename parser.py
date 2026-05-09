import math
import os
import re
import time
import xml.etree.ElementTree as ET
from io import BytesIO
from network import build_network
from transformer_rules import is_transformer_point_candidate, text_contains_transformer_pattern


DEFAULT_SOURCE_IDENTIFIERS = ["DCC7", "DCC", "TAL0001"]
DEFAULT_SOURCE_COORDINATES = (15.59822, 120.92152)
VALIDATION_SUMMARY_KEYS = [
    "total_nodes",
    "total_edges",
    "total_accounts",
    "duplicate_towers",
    "duplicate_accounts",
    "missing_coordinates",
    "unmatched_accounts",
    "invalid_kml_features",
    "disconnected_nodes",
    "inferred_edges",
    "manual_overrides_applied",
    "missing_kwhr_rows",
    "invalid_kwhr_rows",
    "duplicate_frombus_ids",
    "duplicate_tobus_ids",
]


POLE_POINT_PATTERN = re.compile(r"\b(?:SDO|TAL|ALG|MNZ|QZN|LPO|GBA)\d+(?:-UB|-OS)?\b", re.IGNORECASE)


def get_source_identifiers(default=None):
    raw_value = str(os.environ.get("OUTAGE_SOURCE_IDENTIFIERS") or "").strip()
    if raw_value:
        return [item.strip().upper() for item in raw_value.split(",") if item.strip()]
    source_items = list(default or DEFAULT_SOURCE_IDENTIFIERS)
    return [str(item).strip().upper() for item in source_items if str(item).strip()]


def get_source_coordinates(default=None):
    raw_value = str(os.environ.get("OUTAGE_SOURCE_COORDINATES") or "").strip()
    if raw_value:
        parts = [part.strip() for part in raw_value.split(",")]
        if len(parts) == 2:
            try:
                return (float(parts[0]), float(parts[1]))
            except (TypeError, ValueError):
                pass
    if default and len(default) == 2:
        return (float(default[0]), float(default[1]))
    return DEFAULT_SOURCE_COORDINATES


class ValidationError(ValueError):
    def __init__(self, message, validation=None):
        self.validation = finalize_validation(validation or make_validation())
        super().__init__(message)


def make_validation():
    return {
        "status": "ok",
        "errors": [],
        "warnings": [],
        "info": [],
        "summary": {key: 0 for key in VALIDATION_SUMMARY_KEYS},
    }


def finalize_validation(validation):
    if validation["errors"]:
        validation["status"] = "error"
    elif validation["warnings"]:
        validation["status"] = "warning"
    else:
        validation["status"] = "ok"
    return validation


def merge_validations(*items):
    merged = make_validation()
    for validation in items:
        if not validation:
            continue
        merged["errors"].extend(validation.get("errors", []))
        merged["warnings"].extend(validation.get("warnings", []))
        merged["info"].extend(validation.get("info", []))
        for key in VALIDATION_SUMMARY_KEYS:
            merged["summary"][key] += int(validation.get("summary", {}).get(key, 0) or 0)
    return finalize_validation(merged)


def count_duplicate_names(points):
    counts = {}
    duplicates = 0
    for point in points:
        name = normalize_id(point.get("name"))
        if not name:
            continue
        counts[name] = counts.get(name, 0) + 1
    for count in counts.values():
        if count > 1:
            duplicates += count - 1
    return duplicates


def kml_color_to_leaflet(value):
    text = str(value or "").strip().lower()
    if len(text) != 8:
        return {"color": "#0b7285", "opacity": 0.9}

    alpha = int(text[0:2], 16) / 255.0
    blue = text[2:4]
    green = text[4:6]
    red = text[6:8]
    return {"color": f"#{red}{green}{blue}", "opacity": round(alpha, 3)}


def parse_kml_style_maps(root):
    styles = {}
    style_maps = {}

    for elem in root.iter():
        tag = local_name(elem.tag)
        style_id = elem.attrib.get("id", "").strip()

        if tag == "style" and style_id:
            line_style = {"color": "#0b7285", "opacity": 0.9, "weight": 3}
            for child in elem.iter():
                child_tag = local_name(child.tag)
                if child_tag == "color" and child.text:
                    line_style.update(kml_color_to_leaflet(child.text))
                elif child_tag == "width" and child.text:
                    try:
                        line_style["weight"] = max(1, float(child.text.strip()))
                    except ValueError:
                        pass
            styles[f"#{style_id}"] = line_style

        if tag == "stylemap" and style_id:
            normal_ref = ""
            current_key = ""
            for child in list(elem):
                child_tag = local_name(child.tag)
                if child_tag != "pair":
                    continue
                current_key = ""
                current_ref = ""
                for pair_child in list(child):
                    pair_tag = local_name(pair_child.tag)
                    if pair_tag == "key" and pair_child.text:
                        current_key = pair_child.text.strip().lower()
                    elif pair_tag == "styleurl" and pair_child.text:
                        current_ref = pair_child.text.strip()
                if current_key == "normal" and current_ref:
                    normal_ref = current_ref
            if normal_ref:
                style_maps[f"#{style_id}"] = normal_ref

    return styles, style_maps


def dmm_to_decimal(degrees, minutes, direction):
    value = float(degrees) + (float(minutes) / 60.0)
    if direction.upper() in ("S", "W"):
        value *= -1
    return value


def parse_uploaded_file(file_storage):
    raw = file_storage.read()
    validation = make_validation()

    gpx_data = parse_gpx_bytes(raw)
    if gpx_data["points"]:
        return gpx_data

    text = raw.decode("latin-1", errors="ignore")
    points = parse_coordinate_text(text)
    if points:
        validation["summary"]["total_nodes"] = len(points)
        validation["summary"]["duplicate_towers"] = count_duplicate_names(points)
        validation["info"].append("Feeder coordinates were read from plain text rows.")
        return {"points": points, "route_edges": [], "validation": finalize_validation(validation)}

    validation = merge_validations(validation, gpx_data.get("validation"))
    if any("could not be parsed" in str(message).lower() for message in validation["errors"]):
        validation["errors"].append("The feeder GPX appears unreadable or malformed.")
        raise ValidationError(
            "Unreadable GPX file. Please upload a valid GPX export from your mapping source.",
            validation,
        )

    validation["errors"].append("No feeder coordinates were found in the uploaded file.")
    raise ValidationError(
        "No feeder coordinates were found. Check that the GPX contains waypoint or track coordinates, or upload text rows with pole coordinates.",
        validation,
    )


def parse_kml_coordinates_text(text):
    coords = []
    for token in re.split(r"\s+", str(text or "").strip()):
        if not token:
            continue
        parts = [part.strip() for part in token.split(",")]
        if len(parts) < 2:
            continue
        try:
            lon = float(parts[0])
            lat = float(parts[1])
        except ValueError:
            continue
        coords.append([lat, lon])
    return coords


def extract_kml_extended_data_text(placemark):
    chunks = []
    for elem in placemark.iter():
        tag = local_name(elem.tag)
        if tag == "simpledata":
            label = elem.attrib.get("name", "").strip()
            value = (elem.text or "").strip()
            if label or value:
                chunks.append(f"{label} {value}".strip())
        elif tag == "data":
            label = elem.attrib.get("name", "").strip()
            value_text = ""
            for child in list(elem):
                if local_name(child.tag) == "value" and child.text:
                    value_text = child.text.strip()
                    break
            if not value_text and elem.text:
                value_text = elem.text.strip()
            if label or value_text:
                chunks.append(f"{label} {value_text}".strip())
    return " | ".join(chunk for chunk in chunks if chunk)


def extract_tower_like_tokens(*texts):
    tokens = set()
    for text in texts:
        value = str(text or "").strip()
        if not value:
            continue
        normalized_whole = normalize_id(value)
        if normalized_whole:
            tokens.add(normalized_whole)
        for match in POLE_POINT_PATTERN.findall(value):
            normalized_match = normalize_id(match)
            if normalized_match:
                tokens.add(normalized_match)
        for token in re.split(r"[^A-Za-z0-9_-]+", value):
            normalized_token = normalize_id(token)
            if normalized_token:
                tokens.add(normalized_token)
    return {token for token in tokens if token}


def get_polygon_coordinates(geometry):
    for child in list(geometry):
        child_tag = local_name(child.tag)
        if child_tag == "outerboundaryis":
            for descendant in child.iter():
                if local_name(descendant.tag) == "coordinates" and descendant.text:
                    return parse_kml_coordinates_text(descendant.text)
        if child_tag == "linearring":
            for descendant in child.iter():
                if local_name(descendant.tag) == "coordinates" and descendant.text:
                    return parse_kml_coordinates_text(descendant.text)
    for descendant in geometry.iter():
        if local_name(descendant.tag) == "coordinates" and descendant.text:
            return parse_kml_coordinates_text(descendant.text)
    return []


def collect_kml_geometry_entries(node, geometry_entries, stats):
    tag = local_name(node.tag)

    if tag == "multigeometry":
        for child in list(node):
            collect_kml_geometry_entries(child, geometry_entries, stats)
        return

    if tag in {"groundoverlay", "networklink", "screenoverlay", "photooverlay"}:
        return

    if "gx:" in str(node.tag).lower():
        return

    if tag == "linestring":
        coords = []
        for descendant in node.iter():
            if local_name(descendant.tag) == "coordinates" and descendant.text:
                coords = parse_kml_coordinates_text(descendant.text)
                break
        if len(coords) < 2:
            stats["invalid_coordinate_blocks"] += 1
            return
        geometry_entries.append(("linestring", coords))
        return

    if tag == "polygon":
        coords = get_polygon_coordinates(node)
        if len(coords) < 2:
            stats["invalid_coordinate_blocks"] += 1
            return
        geometry_entries.append(("polygon", coords))
        return

    if tag == "linearring":
        coords = []
        for descendant in node.iter():
            if local_name(descendant.tag) == "coordinates" and descendant.text:
                coords = parse_kml_coordinates_text(descendant.text)
                break
        if len(coords) < 2:
            stats["invalid_coordinate_blocks"] += 1
            return
        geometry_entries.append(("polygon", coords))
        return

    if tag == "point":
        coords = []
        for descendant in node.iter():
            if local_name(descendant.tag) == "coordinates" and descendant.text:
                coords = parse_kml_coordinates_text(descendant.text)
                break
        if not coords:
            stats["invalid_coordinate_blocks"] += 1
            return
        geometry_entries.append(("point", [coords[0]]))
        return

    if tag in {
        "name", "description", "styleurl", "style", "extendeddata", "schemadata",
        "simpledata", "data", "value", "snippet", "timestamp", "timespan",
        "visibility", "open", "lookat", "stylemap", "pair", "schema"
    }:
        return

    stats["unsupported_geometries"] += 1


def get_feature_link_tokens(feature_name, feature_description, extended_data_text, style_url, resolved_style_url, geometry_type, coords, tower_points):
    tokens = extract_tower_like_tokens(feature_name, feature_description, extended_data_text)
    style_text = f"{style_url} {resolved_style_url}".upper()
    if geometry_type == "point" and "PRIPOLE" in style_text:
        tokens.update(extract_tower_like_tokens(feature_name, feature_description, extended_data_text, style_text))

    if tower_points and coords:
        probe_points = [coords[0]]
        if len(coords) > 1:
            probe_points.append(coords[-1])
        for probe in probe_points:
            best_name = ""
            best_distance = None
            for tower in tower_points:
                try:
                    tower_lat = float(tower.get("lat"))
                    tower_lon = float(tower.get("lon"))
                except (TypeError, ValueError, AttributeError):
                    continue
                distance = distance_sq((probe[0], probe[1]), (tower_lat, tower_lon))
                if best_distance is None or distance < best_distance:
                    best_distance = distance
                    best_name = tower.get("name", "")
            if best_name and best_distance is not None and best_distance <= 0.000006:
                normalized = normalize_id(best_name)
                if normalized:
                    tokens.add(normalized)

    return {token for token in tokens if token}


def parse_kml_overlay_file(file_storage, tower_names=None, tower_points=None):
    raw = file_storage.read()
    validation = make_validation()
    if not raw:
        validation["errors"].append("The KML file is empty.")
        raise ValidationError("The KML file is empty.", validation)

    decoded_variants = []
    try:
        root = ET.fromstring(raw)
    except Exception:
        root = None
        for encoding in ("utf-8", "utf-8-sig", "utf-16", "latin-1"):
            try:
                decoded_variants.append(raw.decode(encoding, errors="ignore"))
                root = ET.fromstring(decoded_variants[-1])
                break
            except Exception:
                continue
        if root is None:
            validation["errors"].append("The KML file could not be parsed.")
            raise ValidationError("Unreadable KML file. Please upload a valid .kml file.", validation)

    styles, style_maps = parse_kml_style_maps(root)
    features = []
    stats = {
        "placemarks_seen": 0,
        "supported_geometries": 0,
        "unsupported_geometries": 0,
        "invalid_coordinate_blocks": 0,
        "unlinked_features": 0,
        "ignored_point_markers": 0,
    }
    unresolved_features = 0
    normalized_tower_names = {normalize_id(name) for name in (tower_names or []) if normalize_id(name)}
    normalized_tower_points = [tower for tower in (tower_points or []) if isinstance(tower, dict)]
    for placemark_index, placemark in enumerate(root.iter(), start=1):
        if local_name(placemark.tag) != "placemark":
            continue

        stats["placemarks_seen"] += 1

        feature_name = ""
        feature_description = ""
        style_url = ""
        extended_data_text = extract_kml_extended_data_text(placemark)
        for child in list(placemark):
            child_tag = local_name(child.tag)
            if child_tag == "name" and child.text:
                feature_name = child.text.strip()
            elif child_tag == "description" and child.text:
                feature_description = child.text.strip()
            elif child_tag == "styleurl" and child.text:
                style_url = child.text.strip()

        resolved_style_url = style_maps.get(style_url, style_url)
        style = dict(styles.get(resolved_style_url, {"color": "#0b7285", "opacity": 0.9, "weight": 3}))
        geometry_entries = []
        for child in list(placemark):
            collect_kml_geometry_entries(child, geometry_entries, stats)

        if not geometry_entries:
            continue

        for geometry_tag, coords in geometry_entries:
            has_transformer_text = text_contains_transformer_pattern(
                feature_name,
                feature_description,
                extended_data_text,
            )
            is_transformer_point = is_transformer_point_candidate(
                feature_name=feature_name,
                feature_description=feature_description,
                extended_data_text=extended_data_text,
                style_url=style_url,
                resolved_style_url=resolved_style_url,
            )
            feature_tokens = get_feature_link_tokens(
                feature_name,
                feature_description,
                extended_data_text,
                style_url,
                resolved_style_url,
                geometry_tag,
                coords,
                normalized_tower_points,
            )

            if geometry_tag == "point":
                point_style_text = f"{style_url} {resolved_style_url}".upper()
                textual_or_link_hint = bool(feature_tokens.intersection(normalized_tower_names)) if normalized_tower_names else bool(feature_tokens)
                has_point_hint = (
                    "PRIPOLE" in point_style_text
                    or is_transformer_point
                    or bool(POLE_POINT_PATTERN.search(feature_name or ""))
                    or bool(POLE_POINT_PATTERN.search(feature_description or ""))
                    or bool(POLE_POINT_PATTERN.search(extended_data_text or ""))
                    or textual_or_link_hint
                )
                if not has_point_hint:
                    stats["ignored_point_markers"] += 1
                    continue

            if normalized_tower_names and not feature_tokens.intersection(normalized_tower_names):
                unresolved_features += 1
                stats["unlinked_features"] += 1

            stats["supported_geometries"] += 1
            features.append({
                "id": f"kml-feature-{placemark_index}-{len(features) + 1}",
                "name": feature_name,
                "description": feature_description,
                "style_url": style_url,
                "resolved_style_url": resolved_style_url,
                "style": style,
                "geometry": geometry_tag,
                "coords": coords,
                "point_count": len(coords),
                "start": coords[0],
                "end": coords[-1],
                "is_transformer_candidate": bool(
                    is_transformer_point
                    or (geometry_tag != "point" and has_transformer_text)
                ),
            })

    if not features:
        validation["errors"].append("No usable KML features were found.")
        if stats["placemarks_seen"] == 0:
            validation["errors"].append("No Placemark elements were found in the KML file.")
        if stats["invalid_coordinate_blocks"] or stats["unsupported_geometries"]:
            validation["warnings"].append("The KML file only contained invalid, unsupported, or non-linkable geometry.")
        validation["info"].append(f"Placemarks seen: {stats['placemarks_seen']}")
        validation["info"].append(f"Supported geometries found: {stats['supported_geometries']}")
        validation["info"].append(f"Unsupported geometries skipped: {stats['unsupported_geometries']}")
        validation["info"].append(f"Invalid coordinate blocks: {stats['invalid_coordinate_blocks']}")
        validation["info"].append(f"Unlinked features: {stats['unlinked_features']}")
        raise ValidationError(
            "Invalid KML geometry. No usable line, point, or polygon geometry was found after parsing Placemark content.",
            validation,
        )

    validation["summary"]["total_nodes"] = len(features)
    validation["summary"]["invalid_kml_features"] = stats["invalid_coordinate_blocks"] + stats["unsupported_geometries"]
    if stats["unsupported_geometries"]:
        validation["warnings"].append(f"{stats['unsupported_geometries']} KML geometry node(s) used unsupported geometry types and were skipped.")
    if stats["invalid_coordinate_blocks"]:
        validation["warnings"].append(f"{stats['invalid_coordinate_blocks']} KML geometry block(s) had invalid coordinates and were skipped.")
    if normalized_tower_names and unresolved_features:
        validation["warnings"].append(f"{unresolved_features} valid KML feature(s) could not be linked to feeder towers and were kept as overlay-only features.")
    if stats["ignored_point_markers"]:
        validation["info"].append(f"Ignored {stats['ignored_point_markers']} point marker(s) that did not look like feeder or pole references.")
    validation["info"].append(f"Placemarks seen: {stats['placemarks_seen']}")
    validation["info"].append(f"Supported geometries found: {stats['supported_geometries']}")
    validation["info"].append(f"Unsupported geometries skipped: {stats['unsupported_geometries']}")
    validation["info"].append(f"Invalid coordinate blocks: {stats['invalid_coordinate_blocks']}")
    validation["info"].append(f"Unlinked features: {stats['unlinked_features']}")
    validation["info"].append(f"Parsed {len(features)} valid KML feature(s).")

    return {"features": features, "feature_count": len(features), "validation": finalize_validation(validation)}


def local_name(tag):
    if not tag:
        return ""
    if "}" in tag:
        return tag.split("}", 1)[1].lower()
    return tag.lower()


def extract_point_fields(elem):
    lat = elem.attrib.get("lat")
    lon = elem.attrib.get("lon")
    if lat is None or lon is None:
        return None

    try:
        latf = float(lat)
        lonf = float(lon)
    except ValueError:
        return None

    name = ""
    code = ""
    sym = ""
    for child in list(elem):
        child_tag = local_name(child.tag)
        if child_tag == "name" and child.text:
            name = child.text.strip()
        elif child_tag in ("cmt", "desc") and child.text and not code:
            code = child.text.strip()
        elif child_tag == "sym" and child.text:
            sym = child.text.strip()

    if not code:
        code = sym

    return {
        "name": name,
        "code": code,
        "lat": latf,
        "lon": lonf,
    }


def point_identity(point):
    return (
        (point.get("name") or "").strip().upper(),
        round(float(point["lat"]), 8),
        round(float(point["lon"]), 8),
    )


def normalize_route_name(value):
    return str(value or "").strip().upper()


def parse_route_name_pair(route_name):
    text = str(route_name or "").strip()
    if not text or " to " not in text.lower():
        return ("", "")

    parts = re.split(r"\s+to\s+", text, maxsplit=1, flags=re.IGNORECASE)
    if len(parts) != 2:
        return ("", "")

    return parts[0].strip(), parts[1].strip()


def parse_gpx_bytes(raw):
    validation = make_validation()
    root = None

    try:
        root = ET.fromstring(raw)
    except Exception:
        pass

    if root is None:
        parsed_text = None
        for encoding in ("utf-8", "utf-8-sig", "utf-16", "latin-1"):
            try:
                candidate_text = raw.decode(encoding, errors="ignore").strip()
                if "<gpx" not in candidate_text.lower():
                    continue
                root = ET.fromstring(candidate_text)
                parsed_text = candidate_text
                break
            except Exception:
                continue
        if root is None:
            decoded_text = parsed_text or raw.decode("latin-1", errors="ignore").strip()
            if "<gpx" not in decoded_text.lower():
                validation["errors"].append("The uploaded file does not contain GPX data.")
                return {"points": [], "route_edges": [], "validation": finalize_validation(validation)}
            validation["errors"].append("The GPX file could not be parsed.")
            return {"points": [], "route_edges": [], "validation": finalize_validation(validation)}

    wpts = []
    rtepts = []
    trkpts = []
    route_edges = []

    for elem in root.iter():
        tag = local_name(elem.tag)
        if tag == "wpt":
            point = extract_point_fields(elem)
            if point:
                wpts.append(point)
        elif tag == "rtept":
            point = extract_point_fields(elem)
            if point:
                rtepts.append(point)
        elif tag == "trkpt":
            point = extract_point_fields(elem)
            if point:
                trkpts.append(point)

    for route_elem in root.iter():
        if local_name(route_elem.tag) != "rte":
            continue

        route_name = ""
        route_points = []
        for child in list(route_elem):
            child_tag = local_name(child.tag)
            if child_tag == "name" and child.text:
                route_name = child.text.strip()
            elif child_tag == "rtept":
                point = extract_point_fields(child)
                if point:
                    route_points.append(point)

        if len(route_points) < 2:
            continue

        parsed_start, parsed_end = parse_route_name_pair(route_name)
        first_point = dict(route_points[0])
        last_point = dict(route_points[-1])
        if parsed_start:
            first_point["name"] = first_point.get("name") or parsed_start
        if parsed_end:
            last_point["name"] = last_point.get("name") or parsed_end

        route_edges.append((first_point, last_point))

    if wpts:
        points = wpts[:]
    elif rtepts:
        points = rtepts[:]
    else:
        points = simplify_points(trkpts)

    for edge_start, edge_end in route_edges:
        points.append(edge_start)
        points.append(edge_end)

    deduped_points = deduplicate_points(points)
    for i, point in enumerate(deduped_points, start=1):
        if not point["name"]:
            point["name"] = f"TAL{i:04d}"

    validation["summary"]["total_nodes"] = len(deduped_points)
    validation["summary"]["duplicate_towers"] = count_duplicate_names(points)
    if not deduped_points:
        validation["errors"].append("No valid feeder coordinates were found in the GPX file.")
    if len(deduped_points) < 2:
        validation["errors"].append("Less than 2 valid feeder nodes were found in the GPX file.")
    elif route_edges:
        validation["info"].append(f"Loaded {len(route_edges)} explicit feeder route connection(s) from GPX.")
    else:
        validation["warnings"].append("No explicit feeder route connections were found in the GPX file.")

    route_edge_keys = []
    seen_edges = set()
    for edge_start, edge_end in route_edges:
        start_key = point_identity(edge_start)
        end_key = point_identity(edge_end)
        if start_key == end_key:
            continue
        edge_key = (start_key, end_key)
        if edge_key in seen_edges:
            continue
        seen_edges.add(edge_key)
        route_edge_keys.append(edge_key)

    return {"points": deduped_points, "route_edges": route_edge_keys, "validation": finalize_validation(validation)}


def simplify_points(points, max_points=3000):
    n = len(points)
    if n <= max_points:
        return points

    step = int(math.ceil(n / max_points))
    sampled = [points[i] for i in range(0, n, step)]

    if sampled and sampled[-1] is not points[-1]:
        sampled.append(points[-1])

    return sampled


def parse_coordinate_text(text):
    points = []

    pattern = re.compile(
        r"(?P<name>(?:SDO|TAL|ALG|MNZ|QZN|LPO|GBA)[0-9A-Za-z\-]+)\s+"
        r"(?:(?P<code>[A-Za-z0-9\/\-]+)\s+)?"
        r"(?P<lat_dir>[NS])\s*(?P<lat_deg>\d{1,3})\s+(?P<lat_min>\d+(?:\.\d+)?)\s+"
        r"(?P<lon_dir>[EW])\s*(?P<lon_deg>\d{1,3})\s+(?P<lon_min>\d+(?:\.\d+)?)",
        re.IGNORECASE,
    )
    decimal_pattern = re.compile(
        r"(?P<name>(?:SDO|TAL|ALG|MNZ|QZN|LPO|GBA)[0-9A-Za-z\-]+)"
        r"(?:\s+(?P<code>[A-Za-z0-9\/\-]+))?"
        r".*?(?P<lat>-?\d{1,3}\.\d+)\s*[, ]\s*(?P<lon>-?\d{1,3}\.\d+)",
        re.IGNORECASE,
    )

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        match = pattern.search(line)
        if match:
            lat = dmm_to_decimal(match.group("lat_deg"), match.group("lat_min"), match.group("lat_dir"))
            lon = dmm_to_decimal(match.group("lon_deg"), match.group("lon_min"), match.group("lon_dir"))
            points.append(
                {
                    "name": (match.group("name") or "").strip(),
                    "code": (match.group("code") or "").strip(),
                    "lat": lat,
                    "lon": lon,
                }
            )
            continue

        decimal_match = decimal_pattern.search(line)
        if not decimal_match:
            continue

        points.append(
            {
                "name": (decimal_match.group("name") or "").strip(),
                "code": (decimal_match.group("code") or "").strip(),
                "lat": float(decimal_match.group("lat")),
                "lon": float(decimal_match.group("lon")),
            }
        )

    return deduplicate_points(points)


def deduplicate_points(points):
    seen = set()
    cleaned = []

    for point in points:
        key = point_identity(point) + ((point.get("code", "").strip().upper(),),)
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(point)

    return cleaned


def find_source_index(points, source_identifiers=None, source_coordinates=None):
    if not points:
        return None, False

    source_tokens = get_source_identifiers(source_identifiers)
    primary_token = source_tokens[0] if source_tokens else "DCC7"
    last_token = source_tokens[-1] if source_tokens else "TAL0001"
    substation_lat, substation_lon = get_source_coordinates(source_coordinates)

    def substation_score(point):
        code = str(point.get("code", "") or "").strip().upper()
        name = str(point.get("name", "") or "").strip().upper()
        lat = float(point.get("lat", 0.0))
        lon = float(point.get("lon", 0.0))
        distance = (lat - substation_lat) ** 2 + (lon - substation_lon) ** 2
        priority = 3

        if primary_token and primary_token in code:
            priority = 0
        elif any(token and token in code for token in source_tokens[1:]):
            priority = 1
        elif last_token and name.endswith(last_token):
            priority = 2

        return (priority, distance)

    best_index = min(range(len(points)), key=lambda idx: substation_score(points[idx]))
    best_point = points[best_index]
    code = str(best_point.get("code", "") or "").strip().upper()
    name = str(best_point.get("name", "") or "").strip().upper()
    explicit = any(token and token in code for token in source_tokens) or (last_token and name.endswith(last_token))
    return best_index, explicit


def distance_sq(a, b):
    return (a["lat"] - b["lat"]) ** 2 + (a["lon"] - b["lon"]) ** 2


def build_branch_edges(points, source_idx):
    n = len(points)
    in_tree = [False] * n
    best_parent = [-1] * n
    best_dist = [float("inf")] * n

    in_tree[source_idx] = True
    source_locked = False

    for i in range(n):
        if i == source_idx:
            continue
        best_parent[i] = source_idx
        best_dist[i] = distance_sq(points[source_idx], points[i])

    edges = []

    for _ in range(n - 1):
        child = -1
        child_dist = float("inf")

        for i in range(n):
            if in_tree[i]:
                continue
            if best_dist[i] < child_dist:
                child = i
                child_dist = best_dist[i]

        if child == -1:
            break

        parent = best_parent[child]
        edges.append((parent, child))
        in_tree[child] = True

        if parent == source_idx:
            source_locked = True

        for i in range(n):
            if in_tree[i]:
                continue

            d = distance_sq(points[child], points[i])

            if source_locked and best_parent[i] == source_idx:
                best_parent[i] = child
                best_dist[i] = d
            elif d < best_dist[i]:
                best_parent[i] = child
                best_dist[i] = d

    return edges


def unique_tower_names(points):
    counts = {}
    updated = []

    for point in points:
        base_name = (point.get("name") or "").strip() or "Tower"
        key = base_name.upper()
        counts[key] = counts.get(key, 0) + 1

        name = base_name if counts[key] == 1 else f"{base_name}-{counts[key]}"

        item = dict(point)
        item["name"] = name
        updated.append(item)

    return updated


def normalize_id(value):
    if value is None:
        return ""

    text = str(value).strip().upper()
    if not text:
        return ""

    text = text.split()[0]

    if "-" in text:
        first = text.split("-", 1)[0]
        if first.startswith("TAL") or first.startswith("ALG") or first.startswith("MNZ"):
            text = first

    text = re.sub(r"-\d+$", "", text)
    return text.strip()


def normalize_header(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def safe_string(value):
    return "" if value is None else str(value).strip()


def safe_kwhr(value):
    if value is None:
        return 0.0, "missing"

    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0, "missing"

    try:
        return float(text), None
    except ValueError:
        return 0.0, "invalid"


def find_header_index(headers, candidates):
    normalized_headers = [normalize_header(h) for h in headers]
    candidate_set = {normalize_header(c) for c in candidates}

    for idx, header in enumerate(normalized_headers):
        if header in candidate_set:
            return idx

    return None


def make_unique_headers(headers):
    resolved = []
    seen = {}

    for idx, header in enumerate(headers, start=1):
        base = str(header or "").strip() or f"Column {idx}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        resolved.append(base if count == 1 else f"{base} ({count})")

    return resolved


def detect_header_row(rows, max_scan=15):
    best_idx = 0
    best_score = -1

    expected = {
        "frombusid", "frombus", "tobusid", "tobus",
        "polid", "poleid", "accountnumber", "accountno", "acctno",
        "consumername", "customername", "address", "serialnumber", "brand", "kwhr", "kwh"
    }

    for i, row in enumerate(rows[:max_scan]):
        cells = [normalize_header(c) for c in row if c is not None and str(c).strip()]
        if not cells:
            continue

        score = 0
        for cell in cells:
            if cell in expected:
                score += 3
            elif "bus" in cell or "pole" in cell or "pol" in cell or "account" in cell or "acct" in cell or "consumer" in cell or "customer" in cell or "kw" in cell:
                score += 1

        if score > best_score:
            best_score = score
            best_idx = i

    return best_idx


def is_effectively_empty_row(row):
    if not row:
        return True
    for cell in row:
        if cell is None:
            continue
        if str(cell).strip():
            return False
    return True


def parse_xlsx_account_file(file_storage, tower_names=None):
    validation = make_validation()
    try:
        from openpyxl import load_workbook
    except Exception:
        raise ValidationError("XLSX support is missing. Run: pip install openpyxl", validation)

    timings = {
        "workbook_open_ms": 0,
        "header_detection_ms": 0,
        "row_processing_ms": 0,
    }

    stream = getattr(file_storage, "stream", None) or file_storage
    try:
        stream.seek(0)
    except Exception:
        pass

    workbook_open_started = time.perf_counter()
    try:
        workbook = load_workbook(filename=stream, read_only=True, data_only=True)
    except Exception:
        validation["errors"].append("The XLSX file could not be read.")
        raise ValidationError("Unreadable XLSX file. Please upload a valid .xlsx file.", validation)
    timings["workbook_open_ms"] = round((time.perf_counter() - workbook_open_started) * 1000, 2)

    if not workbook.worksheets:
        validation["errors"].append("The XLSX workbook does not contain any worksheets.")
        raise ValidationError("Unreadable XLSX file. No worksheets were found.", validation)

    selected_sheet = None
    header_row_idx = None
    headers = []
    from_idx = None
    to_idx = None
    pol_idx = None
    acct_idx = None
    consumer_name_idx = None
    consumer_type_idx = None
    address_idx = None
    serial_idx = None
    brand_idx = None
    kwhr_idx = None

    header_detection_started = time.perf_counter()
    for sheet in workbook.worksheets:
        sampled_rows = []
        row_iterator = sheet.iter_rows(values_only=True)
        for _, row in zip(range(25), row_iterator):
            sampled_rows.append(row)

        if not sampled_rows or len(sampled_rows) < 2:
            continue

        candidate_header_idx = detect_header_row(sampled_rows, max_scan=min(20, len(sampled_rows)))
        if candidate_header_idx >= len(sampled_rows):
            continue

        candidate_headers = [str(cell).strip() if cell is not None else "" for cell in sampled_rows[candidate_header_idx]]

        candidate_from = find_header_index(candidate_headers, ["frombusid", "from bus id", "frombus", "from bus", "from_bus_id", "from pole", "source bus"])
        candidate_to = find_header_index(candidate_headers, ["tobusid", "to bus id", "tobus", "to bus", "to_bus_id", "to pole", "destination bus"])
        candidate_pol = find_header_index(candidate_headers, ["polid", "pol id", "poleid", "pole id", "tower", "towerid", "tower id", "pole", "pole no", "pole number"])
        candidate_acct = find_header_index(candidate_headers, ["accountnumber", "account number", "acct no", "acctno", "accountno", "account #", "account", "consumer account"])
        candidate_consumer_name = find_header_index(candidate_headers, ["consumer name", "consumername", "consumer_name", "name", "customer name", "consumer"])
        candidate_consumer_type = find_header_index(candidate_headers, ["type", "consumer type", "consumertype", "consumer_type", "customer type"])
        candidate_address = find_header_index(candidate_headers, ["address", "service address", "serviceaddress", "consumer address"])
        candidate_serial = find_header_index(candidate_headers, ["serial", "serial no", "serialno", "serial number", "serialnumber", "meter serial"])
        candidate_brand = find_header_index(candidate_headers, ["brand", "meter brand", "meterbrand"])
        candidate_kwhr = find_header_index(candidate_headers, ["kwhr", "kwhr ", "kwhr", "kwh", "kwhrreading", "kwh reading", "kwhr consumed"])

        has_from_to = candidate_from is not None and candidate_to is not None
        has_pol_acct = candidate_pol is not None or candidate_acct is not None

        if has_pol_acct or has_from_to:
            selected_sheet = sheet
            header_row_idx = candidate_header_idx
            headers = candidate_headers
            from_idx = candidate_from
            to_idx = candidate_to
            pol_idx = candidate_pol
            acct_idx = candidate_acct
            consumer_name_idx = candidate_consumer_name
            consumer_type_idx = candidate_consumer_type
            address_idx = candidate_address
            serial_idx = candidate_serial
            brand_idx = candidate_brand
            kwhr_idx = candidate_kwhr
            break
    timings["header_detection_ms"] = round((time.perf_counter() - header_detection_started) * 1000, 2)

    if selected_sheet is None or header_row_idx is None:
        validation["errors"].append("Required XLSX columns were not found.")
        raise ValidationError(
            "Header row not found or required XLSX columns are missing. Required columns must include Pol ID/Account Number or FromBusID/ToBusID.",
            validation,
        )

    records = []
    append_record = records.append
    normalize_lookup = normalize_id
    to_safe_string = safe_string
    resolved_headers = make_unique_headers(headers)
    duplicate_accounts = 0
    duplicate_pols = 0
    missing_value_rows = 0
    unmatched_count = 0
    unmatched_examples = []
    seen_accounts = set()
    seen_pols = set()
    seen_frombus = set()
    seen_tobus = set()
    duplicate_frombus = 0
    duplicate_tobus = 0
    missing_kwhr_rows = 0
    invalid_kwhr_rows = 0
    feeder_tower_names = {normalize_id(name) for name in (tower_names or []) if normalize_id(name)}
    core_indexes = {
        idx for idx in [
            from_idx,
            to_idx,
            pol_idx,
            acct_idx,
            consumer_name_idx,
            consumer_type_idx,
            address_idx,
            serial_idx,
            brand_idx,
            kwhr_idx,
        ] if idx is not None
    }
    extra_field_columns = [
        (idx, header)
        for idx, header in enumerate(resolved_headers)
        if idx not in core_indexes
    ]

    row_processing_started = time.perf_counter()
    for row_number, row in enumerate(selected_sheet.iter_rows(values_only=True), start=0):
        if row_number <= header_row_idx:
            continue
        if is_effectively_empty_row(row):
            continue

        from_raw = row[from_idx] if from_idx is not None and from_idx < len(row) else None
        to_raw = row[to_idx] if to_idx is not None and to_idx < len(row) else None
        pol_raw = row[pol_idx] if pol_idx is not None and pol_idx < len(row) else None
        acct_raw = row[acct_idx] if acct_idx is not None and acct_idx < len(row) else None

        from_bus_id = to_safe_string(from_raw)
        to_bus_id = to_safe_string(to_raw)
        pol_id = to_safe_string(pol_raw) or from_bus_id
        account_number = to_safe_string(acct_raw) or to_bus_id
        consumer_name = to_safe_string(row[consumer_name_idx] if consumer_name_idx is not None and consumer_name_idx < len(row) else None)
        consumer_type = to_safe_string(row[consumer_type_idx] if consumer_type_idx is not None and consumer_type_idx < len(row) else None)
        address = to_safe_string(row[address_idx] if address_idx is not None and address_idx < len(row) else None)
        serial = to_safe_string(row[serial_idx] if serial_idx is not None and serial_idx < len(row) else None)
        brand = to_safe_string(row[brand_idx] if brand_idx is not None and brand_idx < len(row) else None)
        kwhr_value, kwhr_issue = safe_kwhr(row[kwhr_idx] if kwhr_idx is not None and kwhr_idx < len(row) else None)

        if not pol_id and not account_number:
            continue

        if not pol_id or not account_number:
            missing_value_rows += 1

        if kwhr_issue == "missing":
            missing_kwhr_rows += 1
        elif kwhr_issue == "invalid":
            invalid_kwhr_rows += 1

        extra_fields = {}
        for idx, header in extra_field_columns:
            cell_value = row[idx] if idx < len(row) else None
            text_value = "" if cell_value is None else str(cell_value).strip()
            if text_value:
                extra_fields[header] = text_value

        record = {
            "frombus_id": from_bus_id,
            "tobus_id": to_bus_id,
            "pol_id": pol_id,
            "account_number": account_number,
            "consumer_name": consumer_name,
            "consumer_type": consumer_type,
            "address": address,
            "serial": serial,
            "brand": brand,
            "kwhr": kwhr_value,
            "frombus_norm": normalize_lookup(from_bus_id),
            "tobus_norm": normalize_lookup(to_bus_id),
            "pol_norm": normalize_lookup(pol_id),
            "extra_fields": extra_fields,
        }
        append_record(record)

        if account_number:
            normalized_account = account_number.upper()
            if normalized_account in seen_accounts:
                duplicate_accounts += 1
            else:
                seen_accounts.add(normalized_account)
        if record["pol_norm"]:
            if record["pol_norm"] in seen_pols:
                duplicate_pols += 1
            else:
                seen_pols.add(record["pol_norm"])
        if record["frombus_norm"]:
            if record["frombus_norm"] in seen_frombus:
                duplicate_frombus += 1
            else:
                seen_frombus.add(record["frombus_norm"])
        if record["tobus_norm"]:
            if record["tobus_norm"] in seen_tobus:
                duplicate_tobus += 1
            else:
                seen_tobus.add(record["tobus_norm"])

        if feeder_tower_names:
            match_tokens = {record["pol_norm"], record["frombus_norm"], record["tobus_norm"]}
            if not any(token in feeder_tower_names for token in match_tokens if token):
                unmatched_count += 1
                if len(unmatched_examples) < 20:
                    unmatched_examples.append(
                        f"Pol ID {pol_id or '-'} / Account {account_number or '-'} could not be linked to any feeder tower."
                    )
    timings["row_processing_ms"] = round((time.perf_counter() - row_processing_started) * 1000, 2)

    if not records:
        validation["errors"].append("No usable XLSX rows were found.")
        raise ValidationError(
            "Invalid or missing account mapping values. No usable XLSX rows were found for tower/account mapping.",
            validation,
        )

    validation["summary"]["total_accounts"] = len(records)
    validation["summary"]["duplicate_accounts"] = duplicate_accounts
    validation["summary"]["duplicate_towers"] = duplicate_pols
    validation["summary"]["unmatched_accounts"] = unmatched_count if feeder_tower_names else 0
    validation["summary"]["missing_kwhr_rows"] = missing_kwhr_rows
    validation["summary"]["invalid_kwhr_rows"] = invalid_kwhr_rows
    validation["summary"]["duplicate_frombus_ids"] = duplicate_frombus
    validation["summary"]["duplicate_tobus_ids"] = duplicate_tobus
    if duplicate_accounts:
        validation["warnings"].append(f"{duplicate_accounts} duplicate account number(s) were found in the XLSX file.")
    if duplicate_pols:
        validation["warnings"].append(f"{duplicate_pols} duplicate Pol ID value(s) were found in the XLSX file.")
    if duplicate_frombus:
        validation["warnings"].append(f"{duplicate_frombus} duplicate FromBusID value(s) were found in the XLSX file.")
    if duplicate_tobus:
        validation["warnings"].append(f"{duplicate_tobus} duplicate ToBusID/account fallback value(s) were found in the XLSX file.")
    if missing_value_rows:
        validation["warnings"].append(f"{missing_value_rows} row(s) had invalid or missing account mapping values in Pol ID or account number fields.")
    if missing_kwhr_rows:
        validation["warnings"].append(f"{missing_kwhr_rows} row(s) had missing KWHR values and were treated as 0.")
    if invalid_kwhr_rows:
        validation["warnings"].append(f"{invalid_kwhr_rows} row(s) had invalid KWHR values and were treated as 0.")
    if feeder_tower_names and unmatched_examples:
        validation["warnings"].append(f"{unmatched_count} accounts could not be matched to any tower.")
        validation["info"].extend(unmatched_examples[:20])
    validation["info"].append(f"Loaded {len(records)} account row(s) from XLSX.")
    validation["info"].append(f"Workbook open: {timings['workbook_open_ms']} ms.")
    validation["info"].append(f"Header detection: {timings['header_detection_ms']} ms.")
    validation["info"].append(f"Row processing: {timings['row_processing_ms']} ms.")

    return {
        "headers": resolved_headers,
        "row_count": len(records),
        "records": records,
        "timings": timings,
        "validation": finalize_validation(validation),
    }


def build_network_from_points(points, route_edges=None, manual_overrides=None, source_identifiers=None, source_coordinates=None):
    source_idx, source_explicit = find_source_index(
        points,
        source_identifiers=source_identifiers,
        source_coordinates=source_coordinates,
    )
    if source_idx is None:
        source_idx = 0
    return build_network(
        points,
        source_idx=source_idx,
        route_edges=route_edges,
        manual_overrides=manual_overrides,
        source_explicit=source_explicit,
    )


def build_edges_from_routes(points, route_edges):
    identity_to_index = {point_identity(point): idx for idx, point in enumerate(points)}
    adjacency = {idx: set() for idx in range(len(points))}

    for start_key, end_key in route_edges:
        start_idx = identity_to_index.get(start_key)
        end_idx = identity_to_index.get(end_key)
        if start_idx is None or end_idx is None or start_idx == end_idx:
            continue
        adjacency[start_idx].add(end_idx)
        adjacency[end_idx].add(start_idx)

    if not any(adjacency.values()):
        source_idx, _ = find_source_index(points)
        return build_branch_edges(points, source_idx or 0)

    source_idx, _ = find_source_index(points)
    source_idx = source_idx or 0
    visited = {source_idx}
    queue = [source_idx]
    edges = []

    # Build one rooted feeder tree from the source using the route graph.
    while queue:
        current = queue.pop(0)
        neighbors = sorted(
            adjacency[current],
            key=lambda idx: (distance_sq(points[current], points[idx]), idx),
        )
        for neighbor in neighbors:
            if neighbor in visited:
                continue
            visited.add(neighbor)
            queue.append(neighbor)
            edges.append((current, neighbor))

    if len(visited) == len(points):
        return edges

    # If the GPX/KML route graph has isolated fragments, attach each fragment to the
    # existing feeder tree at the closest already-rooted tower, then keep branching inside it.
    remaining = set(range(len(points))) - visited
    while remaining:
        fragment_root = min(
            remaining,
            key=lambda idx: min(distance_sq(points[idx], points[v]) for v in visited),
        )
        parent = min(visited, key=lambda idx: distance_sq(points[idx], points[fragment_root]))
        edges.append((parent, fragment_root))
        visited.add(fragment_root)
        queue = [fragment_root]
        remaining.remove(fragment_root)

        while queue:
            current = queue.pop(0)
            neighbors = sorted(
                adjacency[current],
                key=lambda idx: (distance_sq(points[current], points[idx]), idx),
            )
            for neighbor in neighbors:
                if neighbor in visited:
                    continue
                visited.add(neighbor)
                queue.append(neighbor)
                edges.append((current, neighbor))
                remaining.discard(neighbor)

    return edges
