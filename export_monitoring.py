from io import BytesIO


MONITORING_HEADERS = [
    "SS",
    "Feeder",
    "Selected Pol ID",
    "Affected Area",
    "Start Time",
    "Restored Date",
    "Restored Time",
    "Action Taken",
    "Duration - Minutes",
    "No. of Cust Affected",
    "Cause of Interruption",
    "Remarks",
    "Estimated KWHR Loss",
    "Estimated Revenue Loss",
    "Created By",
    "Created At",
]


def _autosize_columns(sheet):
    from openpyxl.utils import get_column_letter

    for column_index, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        width = max((len(value) for value in values), default=0) + 2
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(width, 12), 44)


def _format_cause(value):
    return " ".join(part.capitalize() for part in str(value or "unknown").split())


def _write_header(sheet, labels, fill):
    from openpyxl.styles import Font

    for column, label in enumerate(labels, start=1):
        cell = sheet.cell(row=1, column=column, value=label)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")


def _append_monitoring_rows(sheet, rows):
    for row in rows or []:
        sheet.append([
            row.get("substation", ""),
            row.get("feeder", ""),
            row.get("selectedPolId", ""),
            row.get("affectedArea", ""),
            row.get("startTime", ""),
            row.get("restoredDate", ""),
            row.get("restoredTime", ""),
            row.get("actionTaken") or row.get("status", ""),
            row.get("durationMinutes", ""),
            row.get("customersAffected", 0),
            _format_cause(row.get("causeOfInterruption")),
            row.get("remarks", ""),
            row.get("estimatedKwhrLoss", ""),
            row.get("estimatedRevenueLoss", ""),
            row.get("createdBy", ""),
            row.get("createdAt", ""),
        ])


def _build_monitoring_sheet(workbook, title, rows, header_fill):
    sheet = workbook.create_sheet(title)
    _write_header(sheet, MONITORING_HEADERS, header_fill)
    _append_monitoring_rows(sheet, rows)
    if sheet.max_row > 1:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
    return sheet


def build_monitoring_workbook(filtered_dashboard, all_dashboard):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, PatternFill
    except Exception as exc:
        raise ValueError("XLSX export support is missing. Run: pip install openpyxl") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill("solid", fgColor="2E7D32")
    _build_monitoring_sheet(workbook, "Filtered Monitoring", filtered_dashboard.get("rows") or [], header_fill)
    _build_monitoring_sheet(workbook, "All Monitoring", all_dashboard.get("rows") or [], header_fill)

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        _autosize_columns(sheet)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
