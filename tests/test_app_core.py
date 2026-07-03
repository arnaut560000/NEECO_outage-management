import importlib
from io import BytesIO
import os
import shutil
import sys
import tempfile
import unittest
from pathlib import Path
from datetime import datetime, timedelta


class OutageAppTestCase(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.mkdtemp(prefix="outage-tests-")
        self.env_backup = {key: os.environ.get(key) for key in self._env_overrides()}
        for key, value in self._env_overrides().items():
            os.environ[key] = value
        self.app_module = self._load_app_module()
        self.client = self.app_module.app.test_client()

    def tearDown(self):
        for module_name in ("app", "config"):
            sys.modules.pop(module_name, None)
        for key, previous_value in self.env_backup.items():
            if previous_value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = previous_value
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _env_overrides(self):
        return {
            "OUTAGE_ENV": "development",
            "OUTAGE_DEBUG": "0",
            "OUTAGE_AUTO_SEED_ADMIN": "0",
            "OUTAGE_SHOW_SEED_CREDENTIALS": "0",
            "OUTAGE_DB_PATH": os.path.join(self.temp_dir, "users.sqlite3"),
            "OUTAGE_BACKUP_DIR": os.path.join(self.temp_dir, "backups"),
            "OUTAGE_WORKSPACE_CACHE_DIR": os.path.join(self.temp_dir, "workspace_cache"),
            "OUTAGE_SECRET_KEY": "test-secret-key-that-is-long-enough",
        }

    def _load_app_module(self):
        for module_name in ("app", "config"):
            sys.modules.pop(module_name, None)
        return importlib.import_module("app")

    def _create_user(self, username, role="operator", password="Password123"):
        return self.app_module.create_user(
            self.app_module.app.config["AUTH_DB_PATH"],
            username=username,
            password=password,
            role=role,
        )

    def _login_via_session(self, user_id, csrf_token="test-csrf-token"):
        with self.client.session_transaction() as session:
            session["user_id"] = user_id
            session["_csrf_token"] = csrf_token
        return csrf_token

    def _extract_csrf_token(self):
        self.client.get("/login")
        with self.client.session_transaction() as session:
            return session.get("_csrf_token")

    def test_login_flow_succeeds_with_valid_csrf(self):
        self._create_user("alice", role="operator", password="Password123")
        csrf_token = self._extract_csrf_token()

        response = self.client.post(
            "/login",
            data={
                "username": "alice",
                "password": "Password123",
                "csrf_token": csrf_token,
            },
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/"))
        with self.client.session_transaction() as session:
            self.assertIn("user_id", session)

    def test_missing_csrf_blocks_workspace_clear(self):
        user = self._create_user("operator-clear", role="operator")
        self._login_via_session(user["id"])

        response = self.client.post("/workspace/current/clear")

        self.assertEqual(response.status_code, 400)
        payload = response.get_json()
        self.assertFalse(payload["success"])
        self.assertIn("Security token missing or invalid", payload["message"])

    def test_operator_cannot_view_audit_logs(self):
        user = self._create_user("operator-audit", role="operator")
        self._login_via_session(user["id"])

        response = self.client.get("/audit-logs")

        self.assertEqual(response.status_code, 403)

    def test_dashboard_and_operations_pages_load_for_logged_in_user(self):
        user = self._create_user("dashboard-user", role="admin")
        self._login_via_session(user["id"])

        dashboard_response = self.client.get("/")
        operations_response = self.client.get("/operations")
        mobile_response = self.client.get("/mobile")

        self.assertEqual(dashboard_response.status_code, 200)
        self.assertIn(b"NEECO II - AREA 1 OUTAGE MANAGEMENT SYSTEM", dashboard_response.data)
        self.assertIn(b"Mobile App", dashboard_response.data)
        self.assertEqual(operations_response.status_code, 200)
        self.assertIn(b"Upload .GPX file", operations_response.data)
        self.assertEqual(mobile_response.status_code, 200)
        self.assertIn(b"Interruption Monitoring and Informations", mobile_response.data)

    def test_supervisor_can_view_but_not_manage_audit_logs(self):
        user = self._create_user("supervisor-audit", role="supervisor")
        csrf_token = self._login_via_session(user["id"])

        view_response = self.client.get("/audit-logs")
        manage_response = self.client.post(
            "/audit-logs/prune",
            data={"csrf_token": csrf_token},
        )

        self.assertEqual(view_response.status_code, 200)
        self.assertEqual(manage_response.status_code, 403)
        self.assertFalse(manage_response.get_json()["success"])

    def test_records_delete_requires_delete_saved_records_permission(self):
        record_owner = self._create_user("record-owner", role="operator")
        viewer = self._create_user("record-viewer", role="viewer")
        operator_csrf = self._login_via_session(record_owner["id"])

        create_response = self.client.post(
            "/interruptions",
            json={
                "name": "Test Interruption",
                "start_date": "2026-04-06",
                "start_time": "08:00",
                "end_date": "2026-04-06",
                "end_time": "09:00",
                "target_name": "TAL001",
                "affected_towers": [{"name": "TAL001"}],
                "matched_rows": [{"pol_id": "TAL001", "account_number": "12-3456-7890", "kwhr": 10}],
            },
            headers={"X-CSRF-Token": operator_csrf},
        )
        interruption_id = create_response.get_json()["interruption"]["id"]

        viewer_csrf = self._login_via_session(viewer["id"], csrf_token="viewer-csrf")
        delete_response = self.client.post(
            f"/records/interruptions/{interruption_id}/delete",
            data={"csrf_token": viewer_csrf},
        )

        self.assertEqual(delete_response.status_code, 403)
        self.assertFalse(delete_response.get_json()["success"])

    def test_stale_workspace_metadata_is_repaired_when_cache_file_is_missing(self):
        user = self._create_user("workspace-user", role="operator")
        self.app_module.upsert_user_workspace(
            user["id"],
            feederFileName="sample.gpx",
            network={
                "towers": [{"name": "TAL001"}],
                "lines": [{"from": 0, "to": 0}],
                "validation": {"status": "ok", "summary": {}},
                "is_inferred": False,
            },
            feederValidation={"status": "ok", "summary": {}},
        )

        cache_path = Path(self.app_module._workspace_cache_path(user["id"], "network"))
        cache_path.unlink()

        metadata = self.app_module.get_user_workspace_metadata(user["id"])
        workspace = self.app_module.get_user_workspace(user["id"], include_payload=True)

        self.assertIsNone(metadata["network"])
        self.assertIsNone(workspace["network"])
        with self.app_module.get_app_db_connection() as connection:
            row = connection.execute(
                "SELECT network_json, feeder_validation_json FROM user_workspace WHERE user_id = ?",
                (user["id"],),
            ).fetchone()
        self.assertIsNone(row["network_json"])
        self.assertIsNone(row["feeder_validation_json"])

    def test_interruption_create_list_delete_and_export_flow(self):
        user = self._create_user("operator-interrupt", role="operator")
        csrf_token = self._login_via_session(user["id"])

        create_response = self.client.post(
            "/interruptions",
            json={
                "name": "Primary Outage",
                "start_date": "2026-04-06",
                "start_time": "08:00",
                "end_date": "2026-04-06",
                "end_time": "09:15",
                "target_name": "TAL001",
                "source_tower_clicked": "TAL000",
                "affected_towers": [{"name": "TAL001"}],
                "matched_rows": [
                    {
                        "pol_id": "TAL001",
                        "account_number": "12-3456-7890",
                        "consumer_name": "Test Consumer",
                        "kwhr": 12.5,
                    }
                ],
            },
            headers={"X-CSRF-Token": csrf_token},
        )

        self.assertEqual(create_response.status_code, 200)
        created_payload = create_response.get_json()
        self.assertTrue(created_payload["success"])
        interruption_id = created_payload["interruption"]["id"]

        list_response = self.client.get("/interruptions")
        self.assertEqual(list_response.status_code, 200)
        self.assertEqual(len(list_response.get_json()["interruptions"]), 1)

        export_response = self.client.get(f"/interruptions/{interruption_id}/export")
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response.headers["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        delete_response = self.client.delete(
            f"/interruptions/{interruption_id}",
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(delete_response.status_code, 200)
        self.assertTrue(delete_response.get_json()["success"])

    def test_interruption_monitoring_fields_can_be_updated(self):
        user = self._create_user("monitoring-admin", role="admin")
        csrf_token = self._login_via_session(user["id"])

        create_response = self.client.post(
            "/interruptions",
            json={
                "name": "Monitoring Outage",
                "start_date": "2026-05-13",
                "start_time": "08:00",
                "end_date": "2026-05-13",
                "end_time": "09:10",
                "target_name": "TAL001",
                "feeder_name": "F11 TAL.gpx",
                "affected_towers": [{"name": "TAL001"}],
                "matched_rows": [{
                    "pol_id": "TAL001",
                    "account_number": "12-3456-7890",
                    "address": "Purok 3, Barangay San Miguel, Talavera",
                    "kwhr": 24,
                }],
            },
            headers={"X-CSRF-Token": csrf_token},
        )
        interruption_id = create_response.get_json()["interruption"]["id"]

        update_response = self.client.patch(
            f"/interruptions/{interruption_id}/monitoring",
            json={
                "status": "restored",
                "action_taken": "Restored primary line",
                "restored_date": "2026-05-13",
                "restored_time": "09:10",
                "cause_of_interruption": "trees",
                "remarks": "Crew completed restoration.",
            },
            headers={"X-CSRF-Token": csrf_token},
        )

        self.assertEqual(update_response.status_code, 200)
        payload = update_response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["interruption"]["status"], "restored")
        self.assertEqual(payload["interruption"]["actionTaken"], "Restored primary line")
        self.assertEqual(payload["interruption"]["causeOfInterruption"], "trees")
        self.assertEqual(payload["interruption"]["restoredDate"], "2026-05-13")
        self.assertEqual(payload["interruption"]["restoredTime"], "09:10")
        self.assertEqual(payload["interruption"]["remarks"], "Crew completed restoration.")
        self.assertEqual(payload["dashboard"]["counters"]["restored"], 1)
        self.assertEqual(payload["dashboard"]["rows"][0]["affectedArea"], "San Miguel")
        self.assertEqual(payload["dashboard"]["rows"][0]["actionTaken"], "Restored primary line")
        self.assertEqual(payload["dashboard"]["rows"][0]["causeOfInterruption"], "trees")
        self.assertEqual(payload["dashboard"]["rows"][0]["selectedPolId"], "TAL001")
        self.assertEqual(payload["dashboard"]["rows"][0]["restoredDate"], "2026-05-13")
        self.assertEqual(payload["dashboard"]["rows"][0]["restoredTime"], "09:10")
        self.assertIsInstance(payload["dashboard"]["rows"][0]["durationMinutes"], int)

    def test_active_interruption_does_not_get_restored_defaults(self):
        user = self._create_user("active-interrupt", role="operator")
        csrf_token = self._login_via_session(user["id"])
        active_start = datetime.now() - timedelta(minutes=75)

        response = self.client.post(
            "/interruptions",
            json={
                "name": "Active Outage",
                "start_date": active_start.strftime("%Y-%m-%d"),
                "start_time": active_start.strftime("%H:%M"),
                "status": "active",
                "target_name": "TAL001",
                "affected_towers": [{"name": "TAL001"}],
                "matched_rows": [{"pol_id": "TAL001", "account_number": "12-3456-7890", "kwhr": 24}],
            },
            headers={"X-CSRF-Token": csrf_token},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["interruption"]["status"], "active")
        self.assertEqual(payload["interruption"]["restoredDate"], "")
        dashboard = self.app_module.build_dashboard_model()
        self.assertGreaterEqual(dashboard["rows"][0]["durationMinutes"], 74)
        self.assertEqual(dashboard["rows"][0]["restoredDate"], "")
        self.assertNotEqual(dashboard["rows"][0]["estimatedKwhrLoss"], "0")
        self.assertNotEqual(dashboard["rows"][0]["estimatedRevenueLoss"], "0")

    def test_dashboard_loss_metrics_are_duration_minutes_based(self):
        start = datetime(2026, 5, 13, 8, 0)
        rows = [{"kwhr": 24}]

        one_minute = self.app_module._duration_loss_metrics(start, 1, rows)
        sixty_minutes = self.app_module._duration_loss_metrics(start, 60, rows)
        expected_one_hour_kwhr = (24 / 31) / 24
        expected_one_hour_revenue = expected_one_hour_kwhr * 2.0148

        self.assertAlmostEqual(one_minute["kwhr_loss"], round(expected_one_hour_kwhr / 60, 4), places=4)
        self.assertAlmostEqual(sixty_minutes["kwhr_loss"], round(expected_one_hour_kwhr, 4), places=4)
        self.assertAlmostEqual(sixty_minutes["revenue_loss"], round(expected_one_hour_revenue, 4), places=4)

    def test_due_scheduled_interruption_counts_as_active(self):
        user = self._create_user("scheduled-user", role="operator")
        csrf_token = self._login_via_session(user["id"])
        due_start = datetime.now() - timedelta(minutes=5)

        response = self.client.post(
            "/interruptions",
            json={
                "name": "Due Scheduled Outage",
                "start_date": due_start.strftime("%Y-%m-%d"),
                "start_time": due_start.strftime("%H:%M"),
                "status": "scheduled",
                "target_name": "TAL001",
                "affected_towers": [{"name": "TAL001"}],
                "matched_rows": [{"pol_id": "TAL001", "account_number": "12-3456-7890", "kwhr": 24}],
            },
            headers={"X-CSRF-Token": csrf_token},
        )

        self.assertEqual(response.status_code, 200)
        dashboard = self.app_module.build_dashboard_model()
        self.assertEqual(dashboard["counters"]["active"], 1)
        self.assertEqual(dashboard["counters"]["scheduled"], 0)
        self.assertEqual(dashboard["rows"][0]["status"], "active")

    def test_mobile_field_report_creates_monitoring_record(self):
        user = self._create_user("mobile-operator", role="operator")
        csrf_token = self._login_via_session(user["id"])
        self.app_module.upsert_user_workspace(
            user["id"],
            feederFileName="F12 SAMPLE.gpx",
            network={
                "towers": [{"name": "F12-001"}],
                "lines": [],
                "validation": {"status": "ok", "summary": {}},
            },
            accountData={
                "headers": ["Pol ID", "Account Number", "Address", "KWHR"],
                "row_count": 1,
                "records": [{
                    "pol_id": "F12-001",
                    "account_number": "12-3456-7890",
                    "address": "Purok 1, Barangay San Francisco, Talavera",
                    "kwhr": 24,
                }],
                "timings": {},
            },
        )
        saved_feeder = self.app_module.save_uploaded_feeder(
            user["id"],
            "F12 SAMPLE.gpx",
            {
                "towers": [{"name": "F12-001"}],
                "lines": [],
                "validation": {"status": "ok", "summary": {}},
            },
            {"status": "ok", "summary": {}},
        )

        workspace_response = self.client.get("/api/mobile/workspace-pol-ids")
        self.assertEqual(workspace_response.status_code, 200)
        workspace_payload = workspace_response.get_json()
        self.assertTrue(workspace_payload["success"])
        self.assertEqual(workspace_payload["workspace"]["feeders"][0]["filename"], "F12 SAMPLE.gpx")

        search_response = self.client.get(f"/uploaded-feeders/{saved_feeder['id']}/search?q=F12-001")
        self.assertEqual(search_response.status_code, 200)
        self.assertEqual(search_response.get_json()["results"][0]["polId"], "F12-001")

        response = self.client.post(
            "/api/mobile/interruptions",
            json={
                "feeder_id": saved_feeder["id"],
                "pol_id": "F12-001",
                "affected_area": "San Francisco",
                "cause_of_interruption": "trees",
                "remarks": "Reported by field team.",
            },
            headers={"X-CSRF-Token": csrf_token},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["interruption"]["sourceTowerClicked"], "F12-001")
        self.assertEqual(payload["interruption"]["causeOfInterruption"], "trees")
        self.assertEqual(payload["interruption"]["totalAffectedAccounts"], 1)
        self.assertEqual(payload["mobile"]["counters"]["active"], 1)
        self.assertEqual(payload["mobile"]["records"][0]["selectedPolId"], "F12-001")
        self.assertEqual(payload["mobile"]["records"][0]["affectedArea"], "San Francisco")
        self.assertEqual(payload["mobile"]["records"][0]["customersAffected"], 1)

    def test_uploaded_feeder_can_be_restored_and_deleted_by_admin(self):
        admin = self._create_user("feeder-admin", role="admin")
        csrf_token = self._login_via_session(admin["id"])
        saved_feeder = self.app_module.save_uploaded_feeder(
            admin["id"],
            "F21 SAMPLE.gpx",
            {
                "towers": [{"name": "F21-001"}, {"name": "F21-002"}],
                "lines": [{"start_index": 0, "end_index": 1}],
                "validation": {"status": "ok", "summary": {}},
            },
            {"status": "ok", "summary": {}},
        )

        list_response = self.client.get("/uploaded-feeders")
        self.assertEqual(list_response.status_code, 200)
        self.assertEqual(list_response.get_json()["feeders"][0]["feederCode"], "F21")

        restore_response = self.client.post(
            f"/uploaded-feeders/{saved_feeder['id']}/restore",
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(restore_response.status_code, 200)
        restore_payload = restore_response.get_json()
        self.assertTrue(restore_payload["success"])
        self.assertEqual(restore_payload["workspace"]["feederFileName"], "F21 SAMPLE.gpx")
        self.assertEqual(restore_payload["workspace"]["network"]["towers"][0]["name"], "F21-001")

        delete_response = self.client.delete(
            f"/uploaded-feeders/{saved_feeder['id']}",
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(delete_response.status_code, 200)
        self.assertEqual(delete_response.get_json()["feeder"]["feederCode"], "F21")
        self.assertEqual(self.client.get("/uploaded-feeders").get_json()["feeders"], [])

    def test_dashboard_data_endpoint_filters_records(self):
        user = self._create_user("dashboard-filter-admin", role="admin")
        csrf_token = self._login_via_session(user["id"])

        self.client.post(
            "/interruptions",
            json={
                "name": "Talavera Active",
                "start_date": "2026-05-13",
                "start_time": "08:00",
                "status": "active",
                "target_name": "TAL001",
                "feeder_name": "F12 TAL FINAL.gpx",
                "affected_towers": [{"name": "TAL001"}],
                "matched_rows": [{"pol_id": "TAL001", "account_number": "12-3456-7890", "kwhr": 24}],
            },
            headers={"X-CSRF-Token": csrf_token},
        )
        self.client.post(
            "/interruptions",
            json={
                "name": "Guimba Restored",
                "start_date": "2026-05-14",
                "start_time": "08:00",
                "status": "restored",
                "restored_date": "2026-05-14",
                "restored_time": "09:00",
                "target_name": "GBA001",
                "feeder_name": "F41 GBA.gpx",
                "affected_towers": [{"name": "GBA001"}],
                "matched_rows": [{"pol_id": "GBA001", "account_number": "12-3456-7891", "kwhr": 12}],
            },
            headers={"X-CSRF-Token": csrf_token},
        )

        response = self.client.get("/dashboard/data?status=active&feeder=F12")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["dashboard"]["counters"]["total"], 1)
        self.assertEqual(payload["dashboard"]["counters"]["active"], 1)
        self.assertEqual(payload["dashboard"]["rows"][0]["name"], "Talavera Active")
        self.assertEqual(payload["dashboard"]["rows"][0]["selectedPolId"], "TAL001")
        self.assertEqual(payload["dashboard"]["rows"][0]["causeOfInterruption"], "unknown")
        self.assertIn("F12", payload["dashboard"]["filterOptions"]["feeders"])
        self.assertIn("Guimba", payload["dashboard"]["filterOptions"]["substations"])

        mobile_response = self.client.get("/api/mobile/interruptions?status=active&search=TAL001")
        self.assertEqual(mobile_response.status_code, 200)
        mobile_payload = mobile_response.get_json()
        self.assertTrue(mobile_payload["success"])
        self.assertEqual(mobile_payload["mobile"]["counters"]["total"], 1)
        self.assertEqual(mobile_payload["mobile"]["records"][0]["selectedPolId"], "TAL001")
        self.assertIn("focus_pol_id=TAL001", mobile_payload["mobile"]["records"][0]["operationsUrl"])

        export_response = self.client.get("/dashboard/export-monitoring?status=active&feeder=F12")
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response.headers["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(export_response.data.startswith(b"PK"))
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(export_response.data), read_only=True)
        self.assertIn("Filtered Monitoring", workbook.sheetnames)
        self.assertIn("All Monitoring", workbook.sheetnames)
        self.assertEqual(workbook["Filtered Monitoring"].max_row, 2)
        self.assertEqual(workbook["All Monitoring"].max_row, 3)

        dashboard_page = self.client.get("/")
        self.assertIn(b"focus_pol_id=TAL001", dashboard_page.data)

    def test_delete_all_interruptions_requires_admin_and_confirmation(self):
        operator = self._create_user("bulk-delete-operator", role="operator")
        operator_csrf = self._login_via_session(operator["id"])

        create_response = self.client.post(
            "/interruptions",
            json={
                "name": "Bulk Delete Candidate",
                "start_date": "2026-05-13",
                "start_time": "08:00",
                "status": "active",
                "target_name": "TAL001",
                "affected_towers": [{"name": "TAL001"}],
                "matched_rows": [{"pol_id": "TAL001", "account_number": "12-3456-7890", "kwhr": 24}],
            },
            headers={"X-CSRF-Token": operator_csrf},
        )
        self.assertEqual(create_response.status_code, 200)

        blocked_response = self.client.post(
            "/interruptions/delete-all",
            json={"confirmation": "DELETE ALL"},
            headers={"X-CSRF-Token": operator_csrf},
        )
        self.assertEqual(blocked_response.status_code, 403)

        admin = self._create_user("bulk-delete-admin", role="admin")
        admin_csrf = self._login_via_session(admin["id"], csrf_token="admin-bulk-delete-csrf")
        rejected_response = self.client.post(
            "/interruptions/delete-all",
            json={"confirmation": "delete all"},
            headers={"X-CSRF-Token": admin_csrf},
        )
        self.assertEqual(rejected_response.status_code, 400)
        self.assertEqual(len(self.client.get("/interruptions").get_json()["interruptions"]), 1)

        delete_response = self.client.post(
            "/interruptions/delete-all",
            json={"confirmation": "DELETE ALL"},
            headers={"X-CSRF-Token": admin_csrf},
        )

        self.assertEqual(delete_response.status_code, 200)
        payload = delete_response.get_json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["deletedRows"], 1)
        self.assertTrue(os.path.exists(payload["backupPath"]))
        self.assertEqual(payload["dashboard"]["counters"]["total"], 0)
        self.assertEqual(len(self.client.get("/interruptions").get_json()["interruptions"]), 0)


if __name__ == "__main__":
    unittest.main()
