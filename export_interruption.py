from io import BytesIO
import re


DETAIL_HEADERS = [
    "#",
    "Affected Pol ID",
    "Pol ID",
    "Account Number",
    "frombusID",
    "tobusID",
    "Consumer Name",
    "Type",
    "Address",
    "Serial",
    "Brand",
    "KWHR",
    "Matched Via",
]


def _autosize_columns(sheet):
    for column_cells in sheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        width = max((len(value) for value in values), default=0) + 2
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 42)


def _normalized_text(value):
    return str(value or "").strip()


def _extract_pol_id(value):
    match = re.search(r"\b(?:SDO|TAL|ALG|MNZ|QZN|LPO|GBA)\d+\b", str(value or "").upper())
    return match.group(0) if match else ""


def _normalize_tower_name(value):
    text = _normalized_text(value).upper()
    if not text:
        return ""
    text = text.split(" ")[0]
    if "-" in text:
        first = text.split("-", 1)[0]
        if first.startswith(("SDO", "TAL", "ALG", "MNZ", "QZN", "LPO", "GBA")):
            text = first
    text = re.sub(r"-\d+$", "", text)
    return text.strip()


def _canonical_pol_id(value):
    return _extract_pol_id(value) or _normalize_tower_name(value)


def _safe_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _unique_account_count(matched_rows):
    return len({
        _normalized_text(row.get("account_number"))
        for row in matched_rows
        if _normalized_text(row.get("account_number"))
    })


def _canonical_pol_id_entries(towers):
    entries = []
    seen = set()
    for tower in towers:
        pole_id = _canonical_pol_id(tower.get("name"))
        if not pole_id or pole_id in seen:
            continue
        seen.add(pole_id)
        entries.append({
            **tower,
            "name": pole_id,
        })
    return entries


def _unique_pole_tower_count(towers):
    return len(_canonical_pol_id_entries(towers))


def _total_kwhr(matched_rows):
    return round(sum(_safe_float(row.get("kwhr"), 0.0) for row in matched_rows), 4)


def _finalize_sheet(sheet):
    if sheet.max_row > 7 and sheet.max_column > 1:
        sheet.freeze_panes = "A8"
        sheet.auto_filter.ref = f"A7:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
    _autosize_columns(sheet)


def _build_detail_row(row_number, tower_name, row):
    return [
        row_number,
        tower_name,
        row.get("pol_id", ""),
        row.get("account_number", ""),
        row.get("frombus_id", ""),
        row.get("tobus_id", ""),
        row.get("consumer_name", ""),
        row.get("consumer_type", ""),
        row.get("address", ""),
        row.get("serial", ""),
        row.get("brand", ""),
        _safe_float(row.get("kwhr"), 0.0),
        row.get("matched_via", ""),
    ]


def build_interruption_workbook(interruption):
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise ValueError("XLSX export support is missing. Run: pip install openpyxl") from exc

    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "Affected Area Details"

    towers = list(interruption.get("affected_towers", []))
    canonical_towers = _canonical_pol_id_entries(towers)
    matched_rows = list(interruption.get("matched_rows", []))
    source_tower = interruption.get("source_tower_clicked", "")
    total_towers = interruption.get("total_affected_towers") or len(canonical_towers)
    total_accounts = interruption.get("total_affected_accounts") or _unique_account_count(matched_rows)
    feeder_name = interruption.get("feeder_name", "")
    total_affected_kwhr = interruption.get("total_affected_kwhr", _total_kwhr(matched_rows))
    outage_duration_minutes = interruption.get("outage_duration_minutes", 0)
    days_in_month = interruption.get("days_in_month", 0)
    dsm_rate = interruption.get("dsm_rate", 2.0148)
    kwhr_loss_per_duration = interruption.get("kwhr_loss_per_duration", 0.0)
    kwhr_loss_php = interruption.get("kwhr_loss_php", 0.0)
    trace_confidence = interruption.get("trace_confidence", "confirmed")
    inferred_nodes_count = interruption.get("inferred_nodes_count") or interruption.get("audit", {}).get("inferred_nodes_count", 0)
    inferred_accounts_count = interruption.get("inferred_accounts_count") or interruption.get("audit", {}).get("inferred_accounts_count", 0)
    validation_warnings = interruption.get("validation_warnings") or []
    disconnected_fragment_count = len(interruption.get("disconnected_fragments") or [])

    detail_sheet.append(["Interruption Name", interruption.get("name", "")])
    detail_sheet.append(["Start Date", interruption.get("start_date", "")])
    detail_sheet.append(["Start Time", interruption.get("start_time", "")])
    detail_sheet.append(["Finish Date", interruption.get("end_date", "")])
    detail_sheet.append(["Finish Time", interruption.get("end_time", "")])
    detail_sheet.append(["Target", interruption.get("target_name", "")])
    detail_sheet.append(["Source Tower Clicked", source_tower])
    detail_sheet.append(["Total Pol ID", total_towers])
    detail_sheet.append(["Total Affected Accounts", total_accounts])
    detail_sheet.append(["User", interruption.get("user", "Pending User")])
    detail_sheet.append(["Feeder Name", feeder_name])
    detail_sheet.append(["Trace Confidence", trace_confidence])
    detail_sheet.append(["Total Affected KWHR", total_affected_kwhr])
    detail_sheet.append(["Outage Duration Minutes", outage_duration_minutes])
    detail_sheet.append(["Days In Month", days_in_month])
    detail_sheet.append(["DSM Rate", dsm_rate])
    detail_sheet.append(["Computed KWHR Loss", kwhr_loss_per_duration])
    detail_sheet.append(["Computed KWHR Loss in PHP", kwhr_loss_php])
    if trace_confidence != "confirmed":
        detail_sheet.append([
            "Trace Advisory",
            f"This outage trace used {trace_confidence} connectivity. Review inferred nodes/accounts before operational use.",
        ])
        detail_sheet.append(["Inferred Nodes", inferred_nodes_count])
        detail_sheet.append(["Inferred Accounts", inferred_accounts_count])
    if disconnected_fragment_count:
        detail_sheet.append(["Disconnected Fragments Found", disconnected_fragment_count])
    if validation_warnings:
        detail_sheet.append(["Validation Warnings", " | ".join(validation_warnings)])
    if interruption.get("kml_feature"):
        detail_sheet.append(["KML Feature", interruption["kml_feature"].get("name", "")])
        detail_sheet.append(["KML Description", interruption["kml_feature"].get("description", "")])
    detail_sheet.append([])
    detail_sheet.append(DETAIL_HEADERS)

    written_rows = 0
    if matched_rows:
        for row in matched_rows:
            written_rows += 1
            detail_sheet.append(_build_detail_row(written_rows, _canonical_pol_id(row.get("pol_id") or row.get("matched_tower", "")), row))
    else:
        for tower in canonical_towers:
            written_rows += 1
            detail_sheet.append([
                written_rows,
                tower.get("name", ""),
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                0.0,
                "",
            ])

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Interruption Name", interruption.get("name", "")])
    summary_sheet.append(["Start Date", interruption.get("start_date", "")])
    summary_sheet.append(["Start Time", interruption.get("start_time", "")])
    summary_sheet.append(["Finish Date", interruption.get("end_date", "")])
    summary_sheet.append(["Finish Time", interruption.get("end_time", "")])
    summary_sheet.append(["Target", interruption.get("target_name", "")])
    summary_sheet.append(["Source Tower Clicked", source_tower])
    summary_sheet.append(["User", interruption.get("user", "Pending User")])
    summary_sheet.append(["Feeder Name", feeder_name])
    summary_sheet.append(["Trace Confidence", trace_confidence])
    summary_sheet.append(["Total Pol ID", total_towers])
    summary_sheet.append(["Total Affected Customers", total_accounts])
    summary_sheet.append(["Total Affected Accounts", total_accounts])
    summary_sheet.append(["Mapped Rows", len(matched_rows)])
    summary_sheet.append(["Total Affected KWHR", total_affected_kwhr])
    summary_sheet.append(["Outage Duration Minutes", outage_duration_minutes])
    summary_sheet.append(["Days In Month", days_in_month])
    summary_sheet.append(["DSM Rate", dsm_rate])
    summary_sheet.append(["Computed KWHR Loss", kwhr_loss_per_duration])
    summary_sheet.append(["Computed KWHR Loss in PHP", kwhr_loss_php])
    if trace_confidence != "confirmed":
        summary_sheet.append([
            "Trace Advisory",
            f"This workbook contains a {trace_confidence} feeder trace. Review inferred nodes/accounts before relying on this export.",
        ])
        summary_sheet.append(["Inferred Nodes", inferred_nodes_count])
        summary_sheet.append(["Inferred Accounts", inferred_accounts_count])
    if disconnected_fragment_count:
        summary_sheet.append(["Disconnected Fragments Found", disconnected_fragment_count])
    if validation_warnings:
        summary_sheet.append(["Validation Warnings", " | ".join(validation_warnings)])
    if interruption.get("kml_feature"):
        summary_sheet.append(["KML Feature", interruption["kml_feature"].get("name", "")])
        summary_sheet.append(["KML Description", interruption["kml_feature"].get("description", "")])

    accounts_sheet = workbook.create_sheet("Mapped Accounts")
    accounts_sheet.append(DETAIL_HEADERS)
    for idx, row in enumerate(matched_rows, start=1):
        accounts_sheet.append(_build_detail_row(idx, _canonical_pol_id(row.get("pol_id") or row.get("matched_tower", "")), row))

    towers_sheet = workbook.create_sheet("Affected Pol ID")
    towers_sheet.append(["#", "Pol ID"])
    for idx, tower in enumerate(canonical_towers, start=1):
        towers_sheet.append([idx, tower.get("name", "")])

    for sheet in workbook.worksheets:
        _finalize_sheet(sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
